# – Запаковать кодом в zip архив несколько разных файлов: pdf, xlsx, csv;
# – Положить его в ресурсы;
# – Реализовать чтение и проверку содержимого каждого файла из архива не распаковывая сам архив
import zipfile
import csv
from openpyxl.reader.excel import load_workbook
from pypdf import PdfReader
from import_os_path import ZIP_DIR


def test_add_xlsx_in_archive():
    with zipfile.ZipFile(ZIP_DIR) as zip_files:
        with zip_files.open('file_xlsx.xlsx') as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=3, column=1).value == 'perfomance testing'


def test_add_pdf_in_archive():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('file_pdf.pdf') as pdf_file:
            reader = PdfReader(pdf_file)
            page = reader.pages[0]
            text = page.extract_text()
            assert 'Docker' in text


def test_add_csv_in_archive():
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('file_csv.csv') as csv_file:
            content = csv_file.read().decode('utf-8')
            csvreader = list(csv.reader(content.splitlines()))
            head_row = csvreader[0]
            assert head_row[1] == 'format'